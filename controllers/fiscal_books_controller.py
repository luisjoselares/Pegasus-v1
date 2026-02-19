import os
from datetime import datetime
from data.conexion import crear_conexion
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --- IMPORTACIONES PARA PDF (REPORTLAB) ---
from reportlab.lib.pagesizes import legal, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

class FiscalBooksController:
    
    @staticmethod
    def generar_excel_libro_ventas(mes, anio, ruta_guardado):
        # (El código de Excel se mantiene intacto)
        conn = crear_conexion()
        if not conn: return False, "Error de conexión a la base de datos."
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT rif, razon_social FROM configuracion LIMIT 1")
            empresa = cursor.fetchone()
            if not empresa: return False, "Configure los datos de la empresa primero."
            
            mes_str = str(mes).zfill(2)
            anio_str = str(anio)
            
            cursor.execute("""
                SELECT 
                    d.fecha, c.cedula_rif, c.nombre, d.nro_documento, d.nro_control, 
                    d.estado, d.total_usd, d.impuesto_iva_usd, d.tasa_cambio_momento, 
                    d.iva_retenido_bs, d.documento_referencia,
                    COALESCE((
                        SELECT SUM(dd.cantidad * dd.precio_unitario_usd) 
                        FROM documento_detalles dd 
                        JOIN productos p ON dd.producto_id = p.id 
                        WHERE dd.documento_id = d.id AND p.es_exento = 1
                    ), 0) as exento_usd
                FROM documentos d
                JOIN clientes c ON d.cliente_id = c.id
                WHERE d.tipo_doc = 'FACTURA' 
                AND strftime('%m', d.fecha) = ? AND strftime('%Y', d.fecha) = ?
                ORDER BY d.fecha ASC, d.id ASC
            """, (mes_str, anio_str))
            
            facturas = cursor.fetchall()
            if not facturas: return False, "No hay facturas registradas en el período seleccionado."

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Ventas {mes_str}-{anio_str}"
            
            bold_font = Font(bold=True)
            header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            ws['A1'] = f"Contribuyente: {empresa['razon_social']}"
            ws['A1'].font = bold_font
            ws['A2'] = f"RIF: {empresa['rif']}"
            ws['A2'].font = bold_font
            ws['A3'] = f"LIBRO DE VENTAS - PERÍODO: {mes_str}/{anio_str}"
            ws['A3'].font = bold_font
            ws['A4'] = "Expresado en Bolívares (Bs.)"
            ws['A4'].font = Font(italic=True)

            columnas = [
                "Nro. Oper.", "Fecha Fact.", "RIF / C.I.", "Razón Social", 
                "Nro. Factura", "Nro. Control", "Tipo Trans.", "Factura Afectada", 
                "Total Venta", "Ventas Exentas", "Base Imponible", 
                "IVA 16%", "IVA Retenido", "Comprobante Ret."
            ]
            ws.append([]) 
            ws.append(columnas)
            
            for col_num in range(1, len(columnas) + 1):
                celda = ws.cell(row=6, column=col_num)
                celda.font = bold_font; celda.fill = header_fill
                celda.alignment = center_align; celda.border = border
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 16

            ws.column_dimensions['D'].width = 30 

            fila_actual = 7
            totales = {'total': 0, 'exento': 0, 'base': 0, 'iva': 0, 'retenido': 0}
            
            for i, fac in enumerate(facturas, start=1):
                tasa = fac['tasa_cambio_momento']
                es_anulada = fac['estado'] == 'ANULADO'
                
                total_bs = 0 if es_anulada else fac['total_usd'] * tasa
                exento_bs = 0 if es_anulada else fac['exento_usd'] * tasa
                iva_bs = 0 if es_anulada else fac['impuesto_iva_usd'] * tasa
                base_bs = 0 if es_anulada else total_bs - exento_bs - iva_bs
                retenido_bs = 0 if es_anulada else fac['iva_retenido_bs']
                
                fecha_formato = datetime.strptime(fac['fecha'], "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y")
                
                fila_datos = [
                    i, fecha_formato, fac['cedula_rif'], fac['nombre'], 
                    fac['nro_documento'].replace("FAC-", ""), fac['nro_control'], 
                    "02-ANU" if es_anulada else "01-REG", "", 
                    round(total_bs, 2), round(exento_bs, 2), round(base_bs, 2),
                    round(iva_bs, 2), round(retenido_bs, 2),
                    fac['documento_referencia'] if fac['documento_referencia'] else ""
                ]
                
                ws.append(fila_datos)
                for col_num in range(1, len(fila_datos) + 1): ws.cell(row=fila_actual, column=col_num).border = border
                for col_num in range(9, 14): ws.cell(row=fila_actual, column=col_num).number_format = '#,##0.00'
                
                totales['total'] += total_bs; totales['exento'] += exento_bs
                totales['base'] += base_bs; totales['iva'] += iva_bs; totales['retenido'] += retenido_bs
                fila_actual += 1

            ws.append(["", "", "", "TOTALES:", "", "", "", "", 
                       totales['total'], totales['exento'], totales['base'], 
                       totales['iva'], totales['retenido'], ""])
                       
            for col_num in range(4, 14):
                celda = ws.cell(row=fila_actual, column=col_num)
                celda.font = bold_font; celda.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                celda.border = border
                if col_num >= 9: celda.number_format = '#,##0.00'

            wb.save(ruta_guardado)
            return True, "Libro de Ventas Excel generado exitosamente."
        except Exception as e:
            return False, f"Error generando Excel: {e}"
        finally:
            conn.close()

    @staticmethod
    def generar_pdf_libro_ventas(mes, anio, ruta_guardado):
        conn = crear_conexion()
        if not conn: return False, "Error de conexión a la base de datos."
        
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT rif, razon_social FROM configuracion LIMIT 1")
            empresa = cursor.fetchone()
            if not empresa: return False, "Configure los datos de la empresa primero."
            
            mes_str = str(mes).zfill(2)
            anio_str = str(anio)
            
            cursor.execute("""
                SELECT 
                    d.fecha, c.cedula_rif, c.nombre, d.nro_documento, d.nro_control, 
                    d.estado, d.total_usd, d.impuesto_iva_usd, d.tasa_cambio_momento, 
                    d.iva_retenido_bs, d.documento_referencia,
                    COALESCE((
                        SELECT SUM(dd.cantidad * dd.precio_unitario_usd) 
                        FROM documento_detalles dd 
                        JOIN productos p ON dd.producto_id = p.id 
                        WHERE dd.documento_id = d.id AND p.es_exento = 1
                    ), 0) as exento_usd
                FROM documentos d
                JOIN clientes c ON d.cliente_id = c.id
                WHERE d.tipo_doc = 'FACTURA' 
                AND strftime('%m', d.fecha) = ? AND strftime('%Y', d.fecha) = ?
                ORDER BY d.fecha ASC, d.id ASC
            """, (mes_str, anio_str))
            
            facturas = cursor.fetchall()
            if not facturas: return False, "No hay facturas registradas en este mes."

            # --- CONFIGURACIÓN DEL DOCUMENTO PDF (Oficio / Horizontal) ---
            doc = SimpleDocTemplate(ruta_guardado, pagesize=landscape(legal),
                                    rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=30)
            elementos = []
            estilos = getSampleStyleSheet()
            
            # --- ENCABEZADO LEGAL ---
            texto_encabezado = f"""
            <b>Contribuyente:</b> {empresa['razon_social']}<br/>
            <b>RIF:</b> {empresa['rif']}<br/>
            <b>LIBRO DE VENTAS - PERÍODO: {mes_str}/{anio_str}</b><br/>
            <i>Expresado en Bolívares (Bs.)</i>
            """
            elementos.append(Paragraph(texto_encabezado, estilos['Normal']))
            elementos.append(Spacer(1, 15))

            # --- DATOS DE LA TABLA ---
            datos_tabla = [[
                "Oper", "Fecha", "RIF/C.I.", "Razón Social", 
                "Factura", "Control", "Tipo", "Fact. Afectada", 
                "Total (Bs)", "Exento (Bs)", "Base Imp. (Bs)", 
                "IVA 16%", "IVA Ret.", "Comprobante"
            ]]
            
            totales = {'total': 0, 'exento': 0, 'base': 0, 'iva': 0, 'retenido': 0}
            
            estilo_celda = ParagraphStyle(name="CeldaNormal", fontSize=7, leading=8)
            
            for i, fac in enumerate(facturas, start=1):
                tasa = fac['tasa_cambio_momento']
                es_anulada = fac['estado'] == 'ANULADO'
                
                total_bs = 0 if es_anulada else fac['total_usd'] * tasa
                exento_bs = 0 if es_anulada else fac['exento_usd'] * tasa
                iva_bs = 0 if es_anulada else fac['impuesto_iva_usd'] * tasa
                base_bs = 0 if es_anulada else total_bs - exento_bs - iva_bs
                retenido_bs = 0 if es_anulada else fac['iva_retenido_bs']
                
                fecha_formato = datetime.strptime(fac['fecha'], "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%y")
                
                # Se usa Paragraph para la razón social para que se ajuste al ancho
                nombre_p = Paragraph(fac['nombre'][:35], estilo_celda) 
                
                fila = [
                    str(i), 
                    fecha_formato, 
                    fac['cedula_rif'], 
                    nombre_p, 
                    fac['nro_documento'].replace("FAC-", ""), 
                    fac['nro_control'], 
                    "02-ANU" if es_anulada else "01-REG", 
                    "", 
                    f"{total_bs:,.2f}", 
                    f"{exento_bs:,.2f}", 
                    f"{base_bs:,.2f}",
                    f"{iva_bs:,.2f}", 
                    f"{retenido_bs:,.2f}",
                    fac['documento_referencia'] if fac['documento_referencia'] else ""
                ]
                datos_tabla.append(fila)
                
                totales['total'] += total_bs; totales['exento'] += exento_bs
                totales['base'] += base_bs; totales['iva'] += iva_bs; totales['retenido'] += retenido_bs

            # Fila de totales
            datos_tabla.append([
                "", "", "", "TOTALES DEL MES:", "", "", "", "", 
                f"{totales['total']:,.2f}", f"{totales['exento']:,.2f}", f"{totales['base']:,.2f}", 
                f"{totales['iva']:,.2f}", f"{totales['retenido']:,.2f}", ""
            ])

            # --- ANCHOS DE COLUMNA OPTIMIZADOS (Suma ~ 940 puntos para Legal Landscape) ---
            anchos = [30, 45, 60, 160, 50, 50, 45, 60, 75, 75, 75, 65, 65, 75]

            tabla = Table(datos_tabla, colWidths=anchos, repeatRows=1)
            
            # --- ESTILO DE LA TABLA ---
            estilo_tabla = TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#E0E0E0')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.black),
                ('ALIGN', (0,0), (-1,0), 'CENTER'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 7), # Letra pequeña para que quepa todo
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                
                # Alinear a la derecha las columnas de dinero
                ('ALIGN', (8,1), (12,-1), 'RIGHT'),
                
                # Estilo de la última fila (Totales)
                ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#FFFFCC')),
                ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ])
            tabla.setStyle(estilo_tabla)
            
            elementos.append(tabla)
            doc.build(elementos)

            return True, "Libro de Ventas PDF generado exitosamente."
            
        except Exception as e:
            return False, f"Error generando PDF: {e}"
        finally:
            conn.close()